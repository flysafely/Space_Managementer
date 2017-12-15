from tkinter import *
from tkinter import ttk
from tkinter.filedialog import *
from tkinter.messagebox import *
from PIL import Image, ImageTk

from openpyxl import Workbook  # 写excel时候用，因为win32com使用多线程的时候会提示 被呼叫方拒绝接收呼叫 的错误
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

from win32com.client import Dispatch
from win32com.client import constants as constants
from oscrypto._win import symmetric

import tkinter
import win32com.client
import win32api
import win32con
import threading
import pythoncom  # 多线程调用COM
import platform
import os
import time
import datetime
import uuid
import urllib.request
import re
import types
import CheckRegister as ckr
import CheckUpdate as cku

global Status_label, ProgressValue, download_ProgressValue, isRegistered, UserName, Company, Department, Software_Name, Version
global BW_Data_KeyTitle, BW_Title_Mapping, Column_Name, Final_Table_Title, FileName_Mapping, Excel_App, All_Sheets_Data_Dict, Transit_Path
global financial_data_StrValue
global management_data_StrValue
global accumulation_data_StrValue
global pricezone_data_StrValue
global budget_data_StrValue
global floor_ignore_option, category_ignore_option, floor_ignore_checkbutton, category_ignore_checkbutton

Version = "3.0"
Software_Name = "sm"

All_Sheets_Data_Dict = {}

Column_Name = ("",  # 占位符 后面的序号从1开始
               "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
               "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
               "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM",
               "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
               "BA", "BB", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM",
               "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ",
               "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL", "CM",
               "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ",
               )

Final_Table_Title = [  # 按照表格要求的字段顺序排列
    "",  # 占位符 后面的序号从1开始
    "组合标识",
    "门店",
    "分公司",
    "楼层",
    "类别",
    "专柜号",
    "专柜",
    "价格带",
    "经营面积",
    "面积占比",
    "2017年度营业收入计划",
    "2017年度销售计划",
    "2017年度毛利额预算",
    "2017预算毛利率",
    "本期销售额",
    "同期销售额",
    "销售额同比",
    "客流量",
    "客单价(元)",
    "本期累计销售额",
    "同期累计销售额",
    "累计本期销售额楼层占比(%)",
    "累计销售额同比",
    "累计销售计划达成率(%)",
    "本期毛利额",
    "同期毛利额",
    "毛利额同比",
    "本期毛利率",
    "同期毛利率",
    "毛利率同比",
    "本期累计毛利额",
    "同期累计毛利额",
    "累计毛利额同比",
    "本期累计毛利率",
    "同期累计毛利率",
    "累计毛利率同比",
    "累计毛利额占比(%)",
    "累计毛利额占比与面积占比差",
    "累计毛利额达成率(%)",
    "本期营业收入",
    "本期累计营业收入",
    "累计销售坪效(元/㎡)",
    "累计销售坪效与楼层平均坪效差"
]


BW_Title_Mapping = {
    "组合标识": "组合标识",
    "门店": "门店",
    "分公司": "分公司",
    "楼层": "楼层",
    "大类": "类别",
    "主营品类": "类别",
    "专柜号": "专柜号",
    "专柜": "专柜",
    "价格带": "价格带",
    "当前经营面积(M2)": "经营面积",

    "营业收入\n预算": "2017年度营业收入计划",
    "销售收入(含税)\n预算": "2017年度销售计划",
    "营业毛利\n预算": "2017年度毛利额预算",
    "营业毛利率(%)\n预算": "2017预算毛利率",

    "销售收入(含税)\n本年实际": "本期销售额",
    "销售收入(含税)\n上年同期": "同期销售额",  # 同期可删除
    "销售收入(含税)\n同比(%)": "销售额同比",

    "当期客流量(人)": "客流量",

    "累计销售收入(含税)\n本年实际": "本期累计销售额",
    "累计销售收入(含税)\n上年同期": "同期累计销售额",  # 同期可删除
    "累计销售收入(含税)\n同比(%)": "累计销售额同比",

    "营业毛利\n本年实际": "本期毛利额",
    "营业毛利\n上年同期": "同期毛利额",  # 同期可删除
    "营业毛利\n同比(%)": "毛利额同比",
    "营业毛利率(%)\n本年实际": "本期毛利率",
    "营业毛利率(%)\n上年同期": "同期毛利率",
    "营业毛利率(%)\n同比(%)": "毛利率同比",

    "累计营业毛利\n本年实际": "本期累计毛利额",
    "累计营业毛利\n上年同期": "同期累计毛利额",
    "累计营业毛利\n同比(%)": "累计毛利额同比",
    "累计营业毛利率(%)\n本年实际": "本期累计毛利率",
    "累计营业毛利率(%)\n上年同期": "同期累计毛利率",
    "累计营业毛利率(%)\n同比(%)": "累计毛利率同比",

    "营业收入\n本年实际": "本期营业收入",  # 可选项
    "累计营业收入\n本年实际": "本期累计营业收入",  # 可选项
}


FileName_Mapping = {
    "财务数据": "fData",
    "经营数据": "mData",
    "累计数据": "aData",
    "价格数据": "pData",
    "预算数据": "bData",
}

Sum_Formula_info = {
    "经营面积": {"字段": ["经营面积"], "Method": "合计"},
    "2017年度营业收入计划": {"字段": ["2017年度营业收入计划"], "Method": "合计"},
    "2017年度销售计划": {"字段": ["2017年度销售计划"], "Method": "合计"},
    "2017年度毛利额预算": {"字段": ["2017年度毛利额预算"], "Method": "合计"},
    "2017预算毛利率": {"字段": ["2017年度毛利额预算", "2017年度营业收入计划"], "Method": "比值"},
    "本期销售额": {"字段": ["本期销售额"], "Method": "合计"},
    "同期销售额": {"字段": ["同期销售额"], "Method": "合计"},
    "销售额同比": {"字段": ["本期销售额", "同期销售额"], "Method": "同比"},
    "客流量": {"字段": ["客流量"], "Method": "合计"},
    "客单价(元)": {"字段": ["本期销售额", "客流量"], "Method": "比值"},
    "本期累计销售额": {"字段": ["本期累计销售额"], "Method": "合计"},
    "同期累计销售额": {"字段": ["同期累计销售额"], "Method": "合计"},
    "累计销售额同比": {"字段": ["本期累计销售额", "同期累计销售额"], "Method": "比值"},
    "累计销售计划达成率(%)": {"字段": ["本期累计销售额", "2017年度销售计划"], "Method": "比值"},
    "本期毛利额": {"字段": ["本期毛利额"], "Method": "合计"},
    "同期毛利额": {"字段": ["同期毛利额"], "Method": "合计"},
    "毛利额同比": {"字段": ["本期毛利额", "同期毛利额"], "Method": "同比"},
    "本期毛利率": {"字段": ["本期毛利额", "本期营业收入"], "Method": "比值"},
    "本期累计毛利额": {"字段": ["本期累计毛利额"], "Method": "合计"},
    "同期累计毛利额": {"字段": ["同期累计毛利额"], "Method": "合计"},
    "累计毛利额同比": {"字段": ["本期累计毛利额", "同期累计毛利额"], "Method": "比值"},
    "本期累计毛利率": {"字段": ["同期累计毛利额", "本期累计营业收入"], "Method": "比值"},
    "本期营业收入": {"字段": ["本期营业收入"], "Method": "合计"},
    "本期累计营业收入": {"字段": ["本期累计营业收入"], "Method": "合计"},
    "累计销售坪效(元/㎡)": {"字段": ["本期累计销售额", "经营面积"], "Method": "比值"},
}


BW_Data_KeyTitle = {  # 每张表的有效字段
    "fData":
    [
        "门店",
        "分公司",
        "楼层",  # 需要向右位移1得到专柜名称
        "大类",
        "专柜号",
        "销售收入(含税)\n上年同期",
        "销售收入(含税)\n本年实际",
        "销售收入(含税)\n同比(%)",
        "营业毛利\n上年同期",
        "营业毛利\n本年实际",
        "营业毛利\n同比(%)",
        "营业收入\n本年实际",
        "营业毛利率(%)\n上年同期",
        "营业毛利率(%)\n本年实际",
        "营业毛利率(%)\n同比(%)"
    ],
    "mData":
    [
        "门店",
        "分公司",
        "楼层",
        "主营品类",
        "专柜号",
        "当前经营面积(M2)",
        "当期客流量(人)"
    ],
    "aData":
    [
        "门店",
        "分公司",
        "楼层",  # 需要向右位移1得到专柜名称
        "大类",
        "专柜号",
        "销售收入(含税)\n上年同期",
        "销售收入(含税)\n本年实际",
        "销售收入(含税)\n同比(%)",
        "营业毛利\n上年同期",
        "营业毛利\n本年实际",
        "营业毛利\n同比(%)",
        "营业收入\n本年实际",
        "营业毛利率(%)\n上年同期",
        "营业毛利率(%)\n本年实际",
        "营业毛利率(%)\n同比(%)"
    ],
    "pData":
    [
        "柜号",
        "数量",
        "售价"
    ],
    "bData":
    [
        "门店",
        "分公司",
        "楼层",  # 需要向右位移1得到专柜名称
        "大类",
        "专柜号",
        "销售收入(含税)\n预算",
        "营业收入\n预算",
        "营业毛利\n预算",
        "营业毛利率(%)\n预算"
    ]
}


def DownLoad(dbnum, dbsize, size):
    global download_ProgressValue

    '''''回调函数 
    dbnum: 已经下载的数据块 
    dbsize: 数据块的大小 
    size: 远程文件的大小 
    '''
    percent = 100.0 * dbnum * dbsize / size
    if percent > 100:
        percent = 100

    download_ProgressValue.set(percent)

def ignore_option_pass(window,option1,option2):
    global floor_ignore_option, category_ignore_option
    window.withdraw()
    floor_ignore_option = option1
    category_ignore_option = option2


def loadview():

    global Status_label, ProgressValue, download_ProgressValue, Software_Name, Version
    global financial_data_StrValue
    global management_data_StrValue
    global accumulation_data_StrValue
    global pricezone_data_StrValue
    global budget_data_StrValue
    global floor_ignore_option, category_ignore_option, floor_ignore_checkbutton, category_ignore_checkbutton

    root = tkinter.Tk()
    root.title('空间管理报表生成器-version:%s' % Version)
    ico = os.getcwd() + r'\sm.ico'
    root.iconbitmap(ico)

    #####################窗口初始化参数#####################
    windows_params = Check_System_Info(root.winfo_screenwidth() // 2 - 187, root.winfo_screenheight() // 2 - 260)
    root.geometry(windows_params["geometry"])
    root.maxsize(windows_params["maxsize-x"], windows_params["maxsize-y"])
    root.minsize(windows_params["maxsize-x"], windows_params["maxsize-y"])
    textwidth = windows_params["textwidth"]
    Buttonwidth = windows_params["Buttonwidth"]
    Progressbarwidth = windows_params["Progressbarwidth"]

    #####################支付二维码#####################
    pay_windows = Toplevel()
    pay_windows.title("购买方式")
    pay_windows.iconbitmap(ico)
    path = os.getcwd() + r'\QR_Code.png'
    tkimg = ImageTk.PhotoImage(file=path)
    topLabel = Label(pay_windows, image=tkimg)
    topLabel.pack()
    pay_windows.withdraw()

    #####################下载进度条#####################
    download_windows = Toplevel()
    download_windows.title("进度...")
    download_windows.iconbitmap(ico)
    download_ProgressValue = DoubleVar()
    download_ProgressValue.set(0.0)
    ttk.Progressbar(download_windows, 
                                        orient="horizontal",
                                        length=Progressbarwidth,
                                        mode="determinate",
                                        variable=download_ProgressValue).grid(column=1,
                                                                                                                              row=1,
                                                                                                                              sticky=W,
                                                                                                                              columnspan=1)
    download_windows.withdraw()

    ####################自定义模式选择#####################
    ignore_option_window = Toplevel()
    ignore_option_window.geometry("201x60+%s+%s" % (root.winfo_screenwidth() // 2 - 100, root.winfo_screenheight() // 2 - 200))
    ignore_option_window.title("模式选择")
    ignore_option_window.iconbitmap(ico)
    floor_ignore_option = IntVar()
    category_ignore_option = IntVar()

    floor_ignore_checkbutton = Checkbutton(ignore_option_window, text="自定义楼层       ",
                             font='微软雅黑 -13',
                             height=1,
                             variable=floor_ignore_option,
                             state='normal')
    floor_ignore_checkbutton.grid(column=1,
                                                                        row=1,
                                                                        sticky=W,
                                                                        columnspan=1)    
    category_ignore_checkbutton = Checkbutton(ignore_option_window, text="自定义类别",
                             font='微软雅黑 -13',
                             height=1,
                             variable=category_ignore_option,
                             state='normal')
    category_ignore_checkbutton.grid(column=2,
                                                                                row=1,
                                                                                sticky=E + N,
                                                                                columnspan=1)

    Button(ignore_option_window, text="确定",
                                                                       width=20,
                                                                       font='微软雅黑 -13 bold',
                                                                       command = lambda : ignore_option_window.withdraw()).grid(column=1,
                                                                                                                                                row=2,
                                                                                                                                                sticky=W + N + S + E,
                                                                                                                                                columnspan=2)      
    ######################################################

    financial_data_StrValue = StringVar()
    management_data_StrValue = StringVar()
    accumulation_data_StrValue = StringVar()
    pricezone_data_StrValue = StringVar()
    range_of_price_StrValue = StringVar()
    budget_data_StrValue = StringVar()

    financial_data_StrValue.set("导入ZBH_051报表(本月数据)")
    management_data_StrValue.set('导入ZBH_042报表(本月数据)')
    accumulation_data_StrValue.set('导入ZBH_051报表(1月至本月数据)')
    pricezone_data_StrValue.set('导入富基<单品销售日报表>(本月数据)')
    range_of_price_StrValue.set('1-100000')
    budget_data_StrValue.set('导入ZBH_051报表(1月至12月数据)')

    TextBox_Dict = {
                    "fData": financial_data_StrValue,
                    "mData": management_data_StrValue,
                    "aData": accumulation_data_StrValue,
                    "pData": pricezone_data_StrValue,
                    "bData": budget_data_StrValue,
                    }

    #####################财务数据部分#####################

    Label(root, text=" 财务数据：",
          font='微软雅黑 -13',
          justify=LEFT).grid(column=1,
                             row=1,
                             sticky=W)

    Entry(root, font='微软雅黑 -11',
          width=textwidth,
          state='readonly',
          textvariable=financial_data_StrValue,
          justify=LEFT).grid(column=2,
                             row=1,
                             sticky=N + S + E + W)

    Button(root, text="✚",
           width=4,
           font='微软雅黑 -12 bold',
           command=lambda: Get_file_path(TextBox_Dict,
                                         range_of_price_StrValue.get().split('-'),
                                         "fData")).grid(column=3,
                                                        row=1,
                                                        sticky=W)

    #####################经营数据部分#####################

    Label(root, text=" 经营数据：",
          font='微软雅黑 -13',
          justify=LEFT).grid(column=1,
                             row=2,
                             sticky=W)

    Entry(root, font='微软雅黑 -11',
          width=textwidth,
          state='readonly',
          textvariable=management_data_StrValue,
          justify=LEFT).grid(column=2,
                             row=2,
                             sticky=N + S + E + W)

    Button(root, text="✚",
           width=4,
           font='微软雅黑 -12 bold',
           command=lambda: Get_file_path(TextBox_Dict,
                                         range_of_price_StrValue.get().split('-'),
                                         "mData")).grid(column=3,
                                                        row=2,
                                                        sticky=W)

    #####################累计数据部分#####################

    Label(root, text=" 累计数据：",
          font='微软雅黑 -13',
          justify=LEFT).grid(
        column=1,
        row=3,
        sticky=W)

    Entry(root, font='微软雅黑 -11',
          width=textwidth,
          state='readonly',
          textvariable=accumulation_data_StrValue,
          justify=LEFT).grid(column=2,
                             row=3,
                             sticky=N + S + E + W)

    Button(root, text="✚",
           width=4,
           font='微软雅黑 -12 bold',
           command=lambda: Get_file_path(TextBox_Dict,
                                         range_of_price_StrValue.get().split('-'),
                                         "aData")).grid(column=3,
                                                        row=3,
                                                        sticky=W)

    #####################价格带数据部分#####################

    Label(root, text=" 价格数据：",
          font='微软雅黑 -13',
          justify=LEFT).grid(column=1,
                             row=4,
                             sticky=W)

    Entry(root, font='微软雅黑 -11',
          width=int(textwidth / 1.2),
          state='readonly',
          textvariable=pricezone_data_StrValue,
          justify=LEFT).grid(column=2,
                             row=4,
                             sticky=N + W + S)

    Entry(root, font='微软雅黑 -12',
          width=int(textwidth / 4.4),
          textvariable=range_of_price_StrValue,
          justify=LEFT).grid(column=2,
                             row=4,
                             sticky=E + N + S)

    Button(root, text="✚",
           width=4,
           font='微软雅黑 -12 bold',
           command=lambda: Get_file_path(TextBox_Dict,
                                         range_of_price_StrValue.get().split('-'),
                                         "pData")).grid(column=3,
                                                        row=4,
                                                        sticky=W)

    #####################预算数据部分#####################

    Label(root, text=" 预算数据：",
          font='微软雅黑 -13',
          justify=LEFT).grid(column=1,
                             row=5,
                             sticky=W)

    Entry(root, font='微软雅黑 -11',
          width=textwidth,
          state='readonly',
          textvariable=budget_data_StrValue,
          justify=LEFT).grid(column=2,
                             row=5,
                             sticky=N + S + E + W)

    Button(root, text="✚",
           width=4,
           font='微软雅黑 -12 bold',
           command=lambda: Get_file_path(TextBox_Dict,
                                         range_of_price_StrValue.get().split('-'),
                                         "bData")).grid(column=3,
                                                        row=5,
                                                        sticky=W)

    #####################进度条#####################
    ProgressValue = DoubleVar()
    ProgressValue.set(0.0)
    ttk.Progressbar(root, orient="horizontal",
                    length=Progressbarwidth,
                    mode="determinate",
                    variable=ProgressValue).grid(column=1,
                                                 row=7,
                                                 sticky=W,
                                                 columnspan=3)

    #####################开始按钮#####################
    Button(root, text="                                          开始生成",
           font='微软雅黑 -11 bold',
           width=Buttonwidth,
           height=1,
           command=lambda: Start_Build(pay_windows,  # 支付窗口
                                       option1.get(),  # 删除选项
                                       option2.get(),  # 自定义排序选项
                                       [financial_data_StrValue.get(),
                                        management_data_StrValue.get(),
                                        accumulation_data_StrValue.get(),
                                        pricezone_data_StrValue.get(),
                                        budget_data_StrValue.get()]
                                       )).grid(column=1,
                                               row=6,
                                               sticky=W,
                                               columnspan=3)

    #####################删除同期选项#####################
    option1 = IntVar()   # 用来获取复选框是否被勾选，通过chVarDis.get()来获取其的状态,其状态值为int类型 勾选为1  未勾选为0
    del_option = Checkbutton(root, text="删除同期  |",
                             font='微软雅黑 -9',
                             height=1,
                             variable=option1,
                             state='normal')
    del_option.select()
    del_option.grid(column=1,
                    row=6,
                    sticky=W,
                    columnspan=2)

    #####################自定义排序选项#####################
    option2 = IntVar()   # 用来获取复选框是否被勾选，通过chVarDis.get()来获取其的状态,其状态值为int类型 勾选为1  未勾选为0
    sort_option = Checkbutton(root, text="自定义排序",
                              font='微软雅黑 -9',
                              height=1,
                              variable=option2,
                              state='normal')
    sort_option.deselect()
    sort_option.grid(column=2,
                     row=6,
                     sticky=W,
                     columnspan=2)

    ########################################################

    Status_label = StringVar()
    Status_label.set("准备开始")

    l5 = Label(root, font='微软雅黑 -11',
               bg='lightgray',
               textvariable=Status_label,
               justify=LEFT).grid(column=1,
                                  row=8,
                                  sticky=N + S + E + W,
                                  columnspan=3)

    ########################################################
    Button(root, text="模式选择",
           font='微软雅黑 -9',
           width=6,
           height=1,
           command=lambda:ignore_option_window.deiconify()).grid(column=1,
                                               row=8,
                                               sticky=W,
                                               columnspan=1)    

    Button(root, text="检查更新",
           font='微软雅黑 -9',
           width=6,
           height=1,
           command=lambda:Add_Thread(cku.check_update("130.130.200.30",
                                        Software_Name, Version, download_windows, DownLoad))).grid(column=3,
                                                                                                   row=8,
                                                                                                   sticky=W,
                                                                                                   columnspan=1)

     ########################################################                                       
    Add_Thread(lambda: Check_registration_Status_label(
        "http://130.130.200.49", "registrationcode.ini", b"1234567890123456"))

    root.mainloop()


class myThread (threading.Thread):

    def __init__(self, functions):
        threading.Thread.__init__(self)
        self.functions = functions
        self.result = object

    def run(self):
        self.functions()

    def get_result(self):
        return self.result


def Add_Thread(function):
    thread = myThread(function)
    thread.setDaemon(True)
    thread.start()
    return thread


def Add_Thread_Without_Start(function):
    thread = myThread(function)
    thread.setDaemon(True)
    return thread


def Check_System_Info(screen_width, screen_height):
    system_info = platform.platform()
    if "Windows-7" in system_info or "Windows-10" in system_info:

        return {"geometry": '376x220+%s+%s' % (screen_width, screen_height),
                "maxsize-x": 376,
                "maxsize-y": 220,
                "textwidth": 43,
                "Buttonwidth": 53,
                "Progressbarwidth": 376
                }
    else:
        return {"geometry": '397x213+%s+%s' % (screen_width, screen_height),
                "maxsize-x": 397,
                "maxsize-y": 213,
                "textwidth": 45,
                "Buttonwidth": 56,
                "Progressbarwidth": 397
                }


def Refresh_Status_label(info, value):
    global Status_label, ProgressValue
    Status_label.set(info)
    ProgressValue.set(value)


def Create_New_Xlsx(Path, FileName):
    pythoncom.CoInitialize()
    FilePath = Path + "\\" + FileName  # 多线程中需要添加该方法
    try:
        Excel_App = win32com.client.gencache.EnsureDispatch(
            'Excel.Application')
        Excel_App.Visible = False
        Excel_App.DisplayAlerts = False
        if Path:
            if os.path.exists(FilePath):
                os.remove(FilePath)
            Excel_App.Workbooks.Add().SaveAs(FilePath, FileFormat=51)
            Excel_Workbook = Excel_App.Workbooks.Open(FilePath)
            Excel_Workbook.Close()
            return FilePath
    except Exception as e:
        raise e


def Open_WorkBook_By_Openpyxl(path):
    try:
        Workbook = load_workbook(path)
        return Workbook

    except Exception as e:
        tkinter.messagebox.showinfo(
            "表格数据有误！", "请检查导出文件%s格式、内容是否正确，亦或是没有选择对应的正确文件！" % path)
        raise


def Open_Sheet_By_Openpyxl(path, offset):  # openpyxl sheet序号是从0开始的
    try:
        Workbook = load_workbook(path)
        return Workbook.worksheets[0 + offset]

    except Exception as e:
        tkinter.messagebox.showinfo(
            "表格数据有误！", "请检查导出文件%s格式、内容是否正确，亦或是没有选择对应的正确文件！" % path)
        raise


def Set_Cell_Value(Sheet, Row, Column, Value):

    Sheet[Column_Name[Column] + str(Row)] = Value


def Get_Cell_Value(Sheet, row, column):

    return Sheet[Column_Name[column] + str(row)].value


def Get_Desktop_Path():  # 获取桌面路径
    key = win32api.RegOpenKey(win32con.HKEY_CURRENT_USER,
                              r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders', 0, win32con.KEY_READ)
    return win32api.RegQueryValueEx(key, 'Desktop')[0]


def Change_xls_to_xlsx(path):
    Excel_App = Open_Excel_By_Win32com(False)
    Excel_Workbook = Open_Workbook_By_Win32com(Excel_App, path)
    Excel_Workbook.SaveAs(path + "x", FileFormat=51)
    os.remove(path)


def Open_Excel_By_Win32com(visible):
    pythoncom.CoInitialize()  # 多线程中需要添加该方法
    try:
        Excel_App = win32com.client.gencache.EnsureDispatch(
            'Excel.Application')
        Excel_App.Visible = visible
        Excel_App.DisplayAlerts = False
        return Excel_App
    except Exception as e:
        Refresh_Status_label(e, 0.0)


def Open_Workbook_By_Win32com(excel, path):
    try:
        workbook = excel.Workbooks.Open(path)
        return workbook
    except Exception as e:
        Refresh_Status_label(e, 0.0)


def Get_file_path(TextBox_Dict, PriceZone_Range, Data_KeyValue_Key):
    global FileName_Mapping, Transit_Path

    default_info = {}
    for (k, v) in TextBox_Dict.items():
        default_info[k] = v.get()
    Choosed_paths = tkinter.filedialog.askopenfilenames()  # 支持多选

    if len(Choosed_paths) > 1:  # 多选时候，自动按照文件名称填入对相应位置
        Transit_Path = os.path.dirname(
            Choosed_paths[0].replace("/", "\\"))  # 返回一个保存文件的路径给后面生成中转数据表用
        for path in Choosed_paths:
            if path[len(path) - 1:len(path)] != "x":
                Change_xls_to_xlsx(path.replace("/", "\\"))
                split_str = path.split("/")
                file_name = split_str[len(split_str) - 1].split(".")[0]
                Choosed_path = (path + "x").replace("/",
                                                    "\\")  # 转换文件格式后，自动获取新文件路径
                if file_name in FileName_Mapping.keys():  # 防止多选的时候，选择了错误名称的文件
                    TextBox_Dict[FileName_Mapping[
                        file_name]].set("正在检查你的数据.请稍后！")
                    Add_Thread(lambda:Check_Sheet_Validity(FileName_Mapping[file_name],
                                                            Choosed_path,
                                                            TextBox_Dict[
                                                            FileName_Mapping[file_name]],
                                                            PriceZone_Range))
                else:
                    tkinter.messagebox.showinfo(
                        "提示！", "%s.xlsx不是有效的数据源文件！" % file_name)
            else:
                split_str = path.split("/")
                file_name = split_str[len(split_str) - 1].split(".")[0]
                Choosed_path = path.replace("/", "\\")
                if file_name in FileName_Mapping.keys():  # 防止多选的时候，选择了错误名称的文件
                    TextBox_Dict[FileName_Mapping[
                        file_name]].set("正在检查你的数据.请稍后！")
                    Add_Thread(lambda:
                               Check_Sheet_Validity(FileName_Mapping[file_name],
                                                    Choosed_path,
                                                    TextBox_Dict[
                                                        FileName_Mapping[file_name]],
                                                    PriceZone_Range))
                else:
                    tkinter.messagebox.showinfo(
                        "提示！", "%s.xlsx不是有效的数据源文件！" % file_name)
    elif len(Choosed_paths) == 1:  # 单选时候，判断此时选择的文件是否正确
        Transit_Path = os.path.dirname(
            Choosed_paths[0].replace("/", "\\"))  # 返回一个保存文件的路径给后面生成中转数据表用
        path = Choosed_paths[0]
        if path[len(path) - 1:len(path)] != "x":
            Change_xls_to_xlsx(path.replace("/", "\\"))
            Choosed_path = (path + "x").replace("/", "\\")  # 转换文件格式后，自动获取新文件路径
            TextBox_Dict[Data_KeyValue_Key].set("正在检查你的数据.请稍后！")
            Add_Thread(lambda:
                       Check_Sheet_Validity(Data_KeyValue_Key,
                                            Choosed_path,
                                            TextBox_Dict[Data_KeyValue_Key],
                                            PriceZone_Range))
        else:
            Choosed_path = path.replace("/", "\\")
            TextBox_Dict[Data_KeyValue_Key].set("正在检查你的数据.请稍后！")
            Add_Thread(lambda:
                       Check_Sheet_Validity(Data_KeyValue_Key,
                                            Choosed_path,
                                            TextBox_Dict[Data_KeyValue_Key],
                                            PriceZone_Range))

    else:
        for (k, v) in TextBox_Dict.items():
            TextBox_Dict[k].set(default_info[k])


def Check_Sheet_Validity(DataName, choose_path, TextBox, PriceZone_Range):
    global BW_Data_KeyTitle, Excel_App
    global floor_ignore_option,category_ignore_option
    Column_Add_Num = 0
    isValid = True
    Sheet_Offset = {
        "fData": 1,
        "mData": 1,
        "aData": 1,
        "pData": 0,
        "bData": 1
    }  # BW导出报表第一个Sheet为隐藏的需要排除

    Excel_WorkBook = Open_WorkBook_By_Openpyxl(choose_path)
    Excel_Sheet = Excel_WorkBook.worksheets[0 + Sheet_Offset[DataName]]

    Location_Dict = Get_Values_Location(Excel_Sheet,
                                        BW_Data_KeyTitle[DataName]
                                        )

    if len(Location_Dict) != 0:
        for key in BW_Data_KeyTitle[DataName]:
            if key not in list(Location_Dict.keys()):
                if key not in ("楼层", "大类", "主营品类"):
                    isValid = False
                    TextBox.set("错误描述:<%s>数据未找到！" % key)
                else:
                    if key == "楼层":
                        if floor_ignore_option.get() == 1:
                            Location_Dict[key] = (Location_Dict["门店"][0],100)
                        else:
                            isValid = False
                            TextBox.set("未找到‘楼层’信息，请检查！")
                            #tkinter.messagebox.askyesno("警告！", "未选择<忽略楼层>模式，但未找到‘楼层’信息，请检查！")
                    if key == "大类" or key == "主营品类":
                        if category_ignore_option.get() == 1:
                            Location_Dict[key] = (Location_Dict["门店"][0],101) 
                        else:
                            isValid = False
                            TextBox.set("未找到‘类别’信息，请检查！")
                            #tkinter.messagebox.askyesno("警告！","未选择<忽略类别>模式，但未找到‘大类’或‘主营品类’信息，请检查！")  



    else:
        isValid = False
        TextBox.set("错误描述:表格导入选择错误!")

    if isValid == True:
        Sheets_Data_Summary(Excel_WorkBook,
                            Excel_Sheet,
                            TextBox,
                            PriceZone_Range,
                            DataName,
                            Location_Dict
                            )


def Sheets_Data_Summary(Workbook, Sheet, TextBox, PriceZone_Range, DataName, Location_Dict):
    global All_Sheets_Data_Dict
    if DataName in ["fData", "mData", "aData", "bData"]:
        All_Sheets_Data_Dict[DataName] = BW_Data_Get(
            Workbook, Sheet, TextBox, DataName, Location_Dict)
    else:
        All_Sheets_Data_Dict[DataName] = FJ_Data_Get(
            Workbook, Sheet, TextBox, PriceZone_Range, DataName, Location_Dict)


def Get_Values_Location(Sheet, Values):
    location_dict = {}
    Used_Row_Count = Sheet.max_row
    Used_Column_Count = Sheet.max_column
    next_value = True
    for value in Values:
        next_value = True
        for column in range(1, Used_Column_Count + 1):
            if next_value == True:
                for row in range(1, 4):
                    if next_value == True:
                        if Sheet.cell(row=row, column=column).value == value:
                            location_dict[value] = (row, column)
                            next_value = False

    return location_dict


def Get_RowNum_Of_Value_In_Area(Workbook, Sheet, Values, Columns):
    rownum_array = []
    Used_Row_Count = Sheet.max_row
    Used_Column_Count = Sheet.max_column
    Target_Column = []
    for c in Columns:
        Target_Column.append(c[1])

    for value in Values:
        for column in Target_Column:
            for row in range(3, Used_Row_Count + 1):
                if Sheet.cell(row=row, column=column).value == value:
                    rownum_array.append(row)
    return list(set(rownum_array))


def BW_Data_Get(Workbook, Sheet, TextBox, DataName, Location_Dict):
    global BW_Title_Mapping  # BW 字段 和 最终报表 字段的映射
    global BW_Data_KeyTitle
    global floor_ignore_option,category_ignore_option, floor_ignore_checkbutton, category_ignore_checkbutton
    ignore_option_value_dict = {"楼层":(floor_ignore_option,floor_ignore_checkbutton),"类别":(category_ignore_option,category_ignore_checkbutton)}
    # 专柜号、楼层、大类需要向右位移1
    Used_Row_Count = Sheet.max_row
    Used_Column_Count = Sheet.max_column
    Anchor_Cell = Location_Dict["专柜号"]  # 定位锚点
    Column_Offset = {
        "fData": 1,
        "mData": 0,
        "aData": 1,
        "bData": 1
    }  # BW导出报表第一个Sheet为隐藏的需要排除
    # 将空间表的字段名称对应的BW字段名称所在列号存储
    Location_Column_Array = []
    if DataName != "aData":
        for title in BW_Data_KeyTitle[DataName]:
            Location_Column_Array.append((BW_Title_Mapping[title], Location_Dict[title][1]))
        # 单独将 专柜 的列号 添加进去，它是 专柜号的列号+1
    else:
        for title in BW_Data_KeyTitle[DataName]:
            if title in ["销售收入(含税)\n上年同期",
                         "销售收入(含税)\n本年实际",
                         "销售收入(含税)\n同比(%)",
                         "营业毛利\n上年同期",
                         "营业毛利\n本年实际",
                         "营业毛利\n同比(%)",
                         "营业收入\n本年实际",
                         "营业毛利率(%)\n上年同期",
                         "营业毛利率(%)\n本年实际",
                         "营业毛利率(%)\n同比(%)"]:
                Location_Column_Array.append(
                    (BW_Title_Mapping["累计" + title], Location_Dict[title][1]))
            else:
                Location_Column_Array.append(
                    (BW_Title_Mapping[title], Location_Dict[title][1]))

    if DataName != "mData":
        Location_Column_Array.insert(5, ("专柜", Location_Column_Array[4][1] + 1))
    # 筛选出所有显示‘结果’或者‘总体结果’的行号,用于下面循环中进行录入排除
        Result_Cell_RowNum = Get_RowNum_Of_Value_In_Area(Workbook,
                                                                                                                     Sheet,
                                                                                                                     [None],
                                                                                                                     [Location_Column_Array[5]]
                                                                                                                     )
    else:
        Result_Cell_RowNum = Get_RowNum_Of_Value_In_Area(Workbook,
                                                                                                                     Sheet,
                                                                                                                     [None,"结果","总体结果"],
                                                                                                                     [Location_Column_Array[4]]
                                                                                                                     )
    #
    Default_Record_Dict = {"门店": None, 
                                                       "分公司": None,
                                                       "楼层": None, 
                                                       "类别": None}  # 记录同一个门店，分公司，楼层，类别的信息

    Key_Title_Array = {"楼层":"@", "类别":"-"}
    Data_Dict = {}
    for row in range(Anchor_Cell[0] + 1, Used_Row_Count + 1):  # 从'专柜号'的下一行开始录入
        if row not in Result_Cell_RowNum:  # 录入非'结果'或'总体结果'
            Detail_Dict = {}
            for column_info in Location_Column_Array:
                if column_info[0] in ["门店", "分公司"]:
                    if Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value != None:

                        Default_Record_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value
                        Detail_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value  # 获取分公司、专柜号、专柜名称 需要Column+1

                    else:

                        Detail_Dict[column_info[0]] = Default_Record_Dict[column_info[0]]

                elif column_info[0] in Key_Title_Array.keys():

                    if ignore_option_value_dict[column_info[0]][0].get() == 1 and Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value == None:
                        shoppe_value = Sheet.cell(row=row,column=Location_Column_Array[4 + Column_Offset[DataName]][1]).value
                        split_value = Key_Title_Array[column_info[0]]

                        if  len(shoppe_value.split(Key_Title_Array[column_info[0]])) > 1:
                            Detail_Dict[column_info[0]] = shoppe_value.split(split_value)[1] if column_info[0] == "类别" else shoppe_value.split(split_value)[1].split(Key_Title_Array["类别"])[0]
                        else:
                            Detail_Dict[column_info[0]] = "未定义"  # 获取分公司、专柜号、专柜名称 需要Column+1

                    elif ignore_option_value_dict[column_info[0]][0].get() == 1 and Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value != None:

                        ignore_option_value_dict[column_info[0]][1].deselect() #选了忽略字段 但是表中有对应字段数据
                        TextBox.set("已经取消忽略'%s'的模式！" %(column_info[0]))
                        Default_Record_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value
                        Detail_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value

                    elif ignore_option_value_dict[column_info[0]][0].get() == 0 and Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value == None:

                        Detail_Dict[column_info[0]] = Default_Record_Dict[column_info[0]]

                    elif ignore_option_value_dict[column_info[0]][0].get() == 0 and Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value != None:

                        Default_Record_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value
                        Detail_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1] + Column_Offset[DataName]).value

                # 因为BW总的 百分比值为 去掉%号后的值，需要转换成小数
                elif "率" in column_info[0] or "同比" in column_info[0]:
                    if Sheet.cell(row=row, column=column_info[1]).value != None and Sheet.cell(row=row, column=column_info[1]).value != "X":
                        Detail_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1]).value / 100
                    else:
                        Detail_Dict[column_info[0]] = 0
                else:
                    if Sheet.cell(row=row, column=column_info[1]).value != None and Sheet.cell(row=row, column=column_info[1]).value != "X":
                        Detail_Dict[column_info[0]] = Sheet.cell(row=row, column=column_info[1]).value
                    else:
                        Detail_Dict[column_info[0]] = 0
            # 添加一个 识别标识用于 不同表格中的数据匹配整合
            if DataName in ["fData", "aData", "bData"]:  # 这3张表 都是051 字段对应的值一样
                Detail_Dict["组合标识"] = Detail_Dict["分公司"] + "*" + Detail_Dict["楼层"] + "*" + Detail_Dict["类别"] + "*" + Detail_Dict["专柜"]
            else:  # 因为经营数据中的 专柜号 下面就是 专柜名称  所以组合成识别标识的时候直接用专柜号 的值
                Detail_Dict["组合标识"] = Detail_Dict["分公司"] + "*" + Detail_Dict["楼层"] + "*" + Detail_Dict["类别"] + "*" + Detail_Dict["专柜号"]
        Data_Dict[Detail_Dict["组合标识"]] = Detail_Dict
    TextBox.set("导入数据正确!")
    return Data_Dict


def FJ_Data_Get(Workbook, Sheet, TextBox, PriceZone_Range, DataName, Location_Dict):
    Used_Row_Count = Sheet.max_row
    Used_Column_Count = Sheet.max_column
    Anchor_Cell = Location_Dict["柜号"]  # 定位锚点
    #判断价格带范围数据是否合法，不合法则设定为默认值
    if len(PriceZone_Range) == 2 :
        if PriceZone_Range[0].replace(".","").isdigit() and PriceZone_Range[1].replace(".","").isdigit():
            if float(PriceZone_Range[0]) < float(PriceZone_Range[1]):
                PriceZone_Min = float(PriceZone_Range[0])
                PriceZone_Max = float(PriceZone_Range[1])
            else:
                PriceZone_Min = 1
                PriceZone_Max = 100000
        else:
            PriceZone_Min = 1
            PriceZone_Max = 100000
    else:
        PriceZone_Min = 1
        PriceZone_Max = 100000           

    Default_Record_Dict = {"柜号": None, "数量": None, "售价": None}
    Interim_Dict = {}
    Name_Array = []
    Price_Array = []

    for row in range(Anchor_Cell[0] + 1, Used_Row_Count + 1):
        if Sheet.cell(row=row, column=Anchor_Cell[1]).value != None:
            if Sheet.cell(row=row, column=Anchor_Cell[1]).value.replace("[", "").replace("] ", "*") != Default_Record_Dict["柜号"]:
                name = Sheet.cell(row=row, column=Anchor_Cell[
                                  1]).value.replace("[", "").split("] ")
                keyvalue = name[1]  # 交换一下专柜号和专柜名的位置
                Interim_Dict[keyvalue] = []
                Default_Record_Dict["柜号"] = Sheet.cell(row=row, column=Anchor_Cell[
                                                       1]).value.replace("[", "").replace("] ", "*")

    for row in range(Anchor_Cell[0] + 1, Used_Row_Count + 1):
        if Sheet.cell(row=row, column=Anchor_Cell[1]).value != None and Sheet.cell(row=row, column=Location_Dict["数量"][1]).value > 0:
            if Sheet.cell(row=row, column=Location_Dict["售价"][1]).value > PriceZone_Min and Sheet.cell(row=row, column=Location_Dict["售价"][1]).value < PriceZone_Max:
                name = Sheet.cell(row=row, column=Anchor_Cell[
                                  1]).value.replace("[", "").split("] ")
                keyvalue = name[1]  # 交换一下专柜号和专柜名的位置
                Interim_Dict[keyvalue].append(Sheet.cell(
                    row=row, column=Location_Dict["售价"][1]).value)

    Data_Dict = {}

    for (k, v) in Interim_Dict.items():
        if len(sorted(v)) != 0:  # 存在部分品牌，本月只有退票，没有销售，避免出现柜号后的价格数组为空的情况
            Data_Dict[k] = str(sorted(v)[0]) + "-" + \
                str(sorted(v)[len(sorted(v)) - 1])

    TextBox.set("导入数据正确!")
    return Data_Dict


def Start_Build(pay_windows, del_option, sort_option, StrValues):
    global isRegistered, All_Sheets_Data_Dict, UserName, Company, Department
    if len(set(StrValues)) != 1:
        tkinter.messagebox.askyesno("警告！", "表格数据有误！请检查！")
    elif len(set(StrValues)) == 1 and list(set(StrValues))[0] == "导入数据正确!":
        if isRegistered == False:
            pay_windows.deiconify()
            if tkinter.messagebox.askyesno("警告！", "软件'未激活'或者'注册码已过期',请扫描左侧二维码！"):
                pass
            else:
                pay_windows.withdraw()
        else:
            if UserName == "管理员":
                Refresh_Status_label("预处理源数据...", 10.0)
                Add_Thread(lambda: Build_Final_Table(Input_Data_To_NewFile(
                    Formula_Data_Supplement(PriceZone_Data_Supplement())), del_option, sort_option))
            else:
                Company_In_Dict = All_Sheets_Data_Dict["fData"][
                    list(All_Sheets_Data_Dict["fData"].keys())[0]]["门店"]
                Department_In_Dict = All_Sheets_Data_Dict["fData"][
                    list(All_Sheets_Data_Dict["fData"].keys())[0]]["分公司"]
                if Company_In_Dict == Company and Department_In_Dict == Department:
                    Refresh_Status_label("预处理源数据...", 10.0)
                    Add_Thread(lambda: Build_Final_Table(Input_Data_To_NewFile(
                        Formula_Data_Supplement(PriceZone_Data_Supplement())), del_option, sort_option))
                else:
                    tkinter.messagebox.showinfo("警告！", "您没有权限生成<%s>:<%s>的空间管理报表！" % (
                        Company_In_Dict, Department_In_Dict))



def PriceZone_Data_Supplement():
    global All_Sheets_Data_Dict
    Data_Name_Array = ["fData", "mData", "pData", "bData"]
    PriceZone_Data_Key_Array = list(All_Sheets_Data_Dict["pData"].keys())
    for keyvalue in All_Sheets_Data_Dict["aData"]:
        for Data_Name in Data_Name_Array:
            if Data_Name != "pData":
                if keyvalue in All_Sheets_Data_Dict[Data_Name].keys():
                    for title in BW_Data_KeyTitle[Data_Name][5:]:
                        All_Sheets_Data_Dict["aData"][keyvalue][BW_Title_Mapping[title]] = All_Sheets_Data_Dict[Data_Name][keyvalue][BW_Title_Mapping[title]]
                else:
                    for title in BW_Data_KeyTitle[Data_Name][5:]:
                        All_Sheets_Data_Dict["aData"][keyvalue][BW_Title_Mapping[title]] = 0
            else:
                Shoppe_Name = keyvalue.split("*")[3]
                if Shoppe_Name in PriceZone_Data_Key_Array:
                    if All_Sheets_Data_Dict["aData"][keyvalue]["本期销售额"] != 0 and All_Sheets_Data_Dict["aData"][keyvalue]["本期销售额"] != "X":
                        All_Sheets_Data_Dict["aData"][keyvalue][
                            "价格带"] = All_Sheets_Data_Dict["pData"][Shoppe_Name]
                    else:
                        All_Sheets_Data_Dict["aData"][keyvalue]["价格带"] = "-"
                else:
                    All_Sheets_Data_Dict["aData"][keyvalue]["价格带"] = "-"
    return All_Sheets_Data_Dict["aData"]


def Formula_Data_Supplement(Data_Dict):
    Dict_With_Formula_Data = {}
    Sum_Row_Number = len(Data_Dict) + 1
    for (k, v) in Data_Dict.items():
        Dict_With_Formula_Data[k] = v
        for (formula_title, formula_content) in Build_Formula_Content(v, Sum_Row_Number).items():
            Dict_With_Formula_Data[k][formula_title] = formula_content
    return Dict_With_Formula_Data


def Build_Formula_Content(Data_Dict, Sum_Row_Number):
    global Column_Name, Final_Table_Title
    formula_contents_dict = {}
    # 面积占比=经营面积/SUMIF($D$2:$D$ROWCOUNT,楼层,$I$2:$I$ROWCOUNT)
    formula_contents_dict["面积占比"] = "=" + str(Data_Dict["经营面积"]) + "/" + "SUMIF(%s2:%s%s,\"%s\",%s2:%s%s)" % (Column_Name[Final_Table_Title.index("楼层")],
                                                                                                              Column_Name[
                                                                                                                  Final_Table_Title.index("楼层")],
                                                                                                              Sum_Row_Number,
                                                                                                              Data_Dict[
                                                                                                                  "楼层"],
                                                                                                              Column_Name[
                                                                                                                  Final_Table_Title.index("经营面积")],
                                                                                                              Column_Name[
                                                                                                                  Final_Table_Title.index("经营面积")],
                                                                                                              Sum_Row_Number
                                                                                                              )
    # 客单价(元)=本期销售额/客流量*10000
    formula_contents_dict["客单价(元)"] = "=" + str(
        Data_Dict["本期销售额"]) + "/" + str(Data_Dict["客流量"]) + "*10000"
    # 累计本期销售额楼层占比(%)=本期累计销售额/SUMIF($D$2:$D$ROWCOUNT,楼层,$T$2:$T$ROWCOUNT)
    formula_contents_dict["累计本期销售额楼层占比(%)"] = "=" + str(Data_Dict["本期累计销售额"]) + "/" + "SUMIF(%s2:%s%s,\"%s\",%s2:%s%s)" % (Column_Name[Final_Table_Title.index("楼层")],
                                                                                                                           Column_Name[
                                                                                                                               Final_Table_Title.index("楼层")],
                                                                                                                           Sum_Row_Number,
                                                                                                                           Data_Dict[
                                                                                                                               "楼层"],
                                                                                                                           Column_Name[
                                                                                                                               Final_Table_Title.index("本期累计销售额")],
                                                                                                                           Column_Name[
                                                                                                                               Final_Table_Title.index("本期累计销售额")],
                                                                                                                           Sum_Row_Number
                                                                                                                           )
    # 累计销售计划达成率(%)=本期累计销售额/2017年度销售计划
    formula_contents_dict["累计销售计划达成率(%)"] = "=" + str(
        Data_Dict["本期累计销售额"]) + "/" + str(Data_Dict["2017年度销售计划"])
    # 累计毛利额占比(%)=本期累计毛利额/SUMIF($D$2:$D$ROWCOUNT,楼层,$AC$2:$AC$ROWCOUNT)
    formula_contents_dict["累计毛利额占比(%)"] = "=" + str(Data_Dict["本期累计毛利额"]) + "/" + "SUMIF(%s2:%s%s,\"%s\",%s2:%s%s)" % (Column_Name[Final_Table_Title.index("楼层")],
                                                                                                                       Column_Name[
                                                                                                                           Final_Table_Title.index("楼层")],
                                                                                                                       Sum_Row_Number,
                                                                                                                       Data_Dict[
                                                                                                                           "楼层"],
                                                                                                                       Column_Name[
                                                                                                                           Final_Table_Title.index("本期累计毛利额")],
                                                                                                                       Column_Name[
                                                                                                                           Final_Table_Title.index("本期累计毛利额")],
                                                                                                                       Sum_Row_Number
                                                                                                                       )
    # 累计毛利额占比与面积占比差=本期累计毛利额/SUMIF($D$2:$D$ROWCOUNT,楼层,$AC$2:$AC$ROWCOUNT)-经营面积/SUMIF($D$2:$D$ROWCOUNT,楼层,$I$2:$I$ROWCOUNT)
    formula_contents_dict["累计毛利额占比与面积占比差"] = "=" + formula_contents_dict[
        "累计毛利额占比(%)"] + "-" + formula_contents_dict["面积占比"].replace("=", "")
    # 累计毛利额达成率(%)=本期累计毛利额/2017年度毛利额预算
    formula_contents_dict["累计毛利额达成率(%)"] = "=" + str(
        Data_Dict["本期累计毛利额"]) + "/" + str(Data_Dict["2017年度毛利额预算"])
    # 累计销售坪效(元/㎡)=本期累计销售额/经营面积*10000
    formula_contents_dict[
        "累计销售坪效(元/㎡)"] = "=" + str(Data_Dict["本期累计销售额"]) + "/" + str(Data_Dict["经营面积"]) + "*10000"
    # 累计销售坪效与楼层平均坪效差=本期累计销售额/经营面积*10000-(SUMIF($D$2:$D$ROWCOUNT,楼层,$T$2:$T$ROWCOUNT)*10000)/SUMIF($D$2:$D$ROWCOUNT,楼层,$J$2:$J$ROWCOUNT))
    formula_contents_dict["累计销售坪效与楼层平均坪效差"] = str(formula_contents_dict["累计销售坪效(元/㎡)"]) + "-" + "SUMIF(%s2:%s%s,\"%s\",%s2:%s%s)" % (Column_Name[Final_Table_Title.index("楼层")],
                                                                                                                                     Column_Name[
                                                                                                                                         Final_Table_Title.index("楼层")],
                                                                                                                                     Sum_Row_Number,
                                                                                                                                     Data_Dict[
                                                                                                                                         "楼层"],
                                                                                                                                     Column_Name[
                                                                                                                                         Final_Table_Title.index("本期累计销售额")],
                                                                                                                                     Column_Name[
                                                                                                                                         Final_Table_Title.index("本期累计销售额")],
                                                                                                                                     Sum_Row_Number
                                                                                                                                     ) + "/" + "SUMIF(%s2:%s%s,\"%s\",%s2:%s%s)" % (Column_Name[Final_Table_Title.index("楼层")],
                                                                                                                                                                                    Column_Name[
                                                                                                                                         Final_Table_Title.index("楼层")],
        Sum_Row_Number,
        Data_Dict[
                                                                                                                                         "楼层"],
        Column_Name[
                                                                                                                                         Final_Table_Title.index("经营面积")],
        Column_Name[
                                                                                                                                         Final_Table_Title.index("经营面积")],
        Sum_Row_Number) + "*10000"

    return formula_contents_dict


def Input_Data_To_NewFile(Data_Dict):
    global Transit_Path
    File_Name = "空间管理%s.xlsx" % str(time.strftime(
        "%Y-%m-%d %H-%M-%S", time.localtime()))
    Temp_File_Path = Create_New_Xlsx(Transit_Path, File_Name)
    Excel_WorkBook = Open_WorkBook_By_Openpyxl(Temp_File_Path)
    Excel_Sheet = Excel_WorkBook.worksheets[0]

    # 将表头字段信息 写入 字典
    Title_Array = []
    for title in Final_Table_Title:
        Title_Array.append(title)

    Data_Dict["字段名"] = Title_Array  # 字段名 是数组类型  后面用类型判断它

    Input_Row_Thread_Array = []
    Row_Num = 1
    for key in Data_Dict:
        if key != "字段名":
            Row_Num += 1
            Input_Row_Thread_Array.append(threading.Thread(
                target=Input_Data_One_Row, args=(Excel_Sheet, Row_Num, Data_Dict[key])))
        else:
            Input_Row_Thread_Array.append(threading.Thread(
                target=Input_Data_One_Row, args=(Excel_Sheet, 1, Data_Dict[key])))

    for thread in Input_Row_Thread_Array:
        thread.start()

    Excel_WorkBook.save(Temp_File_Path)

    return Temp_File_Path


def Input_Data_One_Row(Sheet, Row, Data_Dict):
    # 添加数据
    for title in Final_Table_Title:
        if title != "":
            if isinstance(Data_Dict, dict):
                Sheet.cell(row=Row, column=Final_Table_Title.index(
                    title)).value = Data_Dict[title]
            else:
                Sheet.cell(row=Row, column=Final_Table_Title.index(
                    title)).value = title


def Set_Range_Layout(obj=object, type="Range", Name="微软雅黑", Size=9, Strikethrough=False, Superscript=False, OutlineFont=False, Shadow=False, Underline=False, Orientation=0, AddIndent=False, IndentLevel=0, ShrinkToFit=False, MergeCells=False, RowHeight=14, WrapText=True, ColumnWidth=9, NumberFormatLocal="0.00_ "):

    if type == "Range":
        obj.Font.Name = Name
        obj.Font.Size = Size
        obj.Font.Strikethrough = Strikethrough
        obj.Font.Superscript = Superscript
        obj.Font.OutlineFont = OutlineFont
        obj.Font.Shadow = Shadow
        obj.Font.Underline = Underline
        obj.HorizontalAlignment = constants.xlCenter
        obj.VerticalAlignment = constants.xlCenter
        obj.Orientation = Orientation
        obj.AddIndent = AddIndent
        obj.IndentLevel = IndentLevel
        obj.ShrinkToFit = ShrinkToFit
        obj.ReadingOrder = constants.xlContext
        obj.MergeCells = MergeCells
        obj.NumberFormatLocal = NumberFormatLocal
        obj.WrapText = WrapText
    elif type == "Column":
        obj.WrapText = True
        obj.ColumnWidth = ColumnWidth
        obj.NumberFormatLocal = NumberFormatLocal
    elif type == "Row":
        obj.WrapText = True
        obj.RowHeight = RowHeight
        obj.NumberFormatLocal = NumberFormatLocal


def Prepare_Floor_Sheets(Sheets, RowCount):
    global Final_Table_Title
    Floor_Array = []
    for i in range(2, RowCount):
        if Sheets(1).Cells(i, Final_Table_Title.index("楼层")).Value != None:
            Floor_Array.append(Sheets(1).Cells(
                i, Final_Table_Title.index("楼层")).Value)
    Floor_List = list(set(Floor_Array))

    if Sheets.Count - (len(Floor_List) + 1) < 0:
        for i in range(abs(Sheets.Count - (len(Floor_List) + 1))):
            Sheets.Add(After=Sheets(Sheets.Count))

    Floor_List.insert(0, "合计")
    Sheet_Num = 0
    for floor in Floor_List:
        Sheet_Num += 1
        Sheets(Sheet_Num).Name = floor


def Get_Max_Row_Num(Sheet, offset):
    Num = 0
    while Sheet.Cells(Num + 1, 1).Value != None:
        Num += 1
    return Num + offset


def Get_Max_Column_Num(Sheet, offset):
    Num = 0
    while Sheet.Cells(1, Num + 1).Value != None:
        Num += 1
    return Num + offset


def Build_Final_Table(FilePath, del_option, sort_option):
    global Column_Name, Final_Table_Title, Sum_Formula_info
    Refresh_Status_label("整理数据内容...", 30.0)
    # 用win32com打开源数据表格
    Excel_App = Open_Excel_By_Win32com(False)
    Excel_Workbook = Open_Workbook_By_Win32com(Excel_App, FilePath)
    Excel_Sheets = Excel_App.Sheets

    All_Cells_Range = "A1:%s%s" % (Column_Name[Get_Max_Column_Num(
        Excel_Sheets(1), 0)], str(Get_Max_Row_Num(Excel_Sheets(1), 1)))
    # 复制粘贴数据 只要值
    Excel_Sheets(1).Range(All_Cells_Range).Copy()
    Excel_Sheets(1).Range(All_Cells_Range).PasteSpecial(Paste=-4163,
                                                        Operation=-4142,
                                                        SkipBlanks=False,
                                                        Transpose=False
                                                        )
    # 替换#DIV/0!为-
    Excel_App.Sheets(1).Range(All_Cells_Range).Replace(What="#DIV/0!",
                                                       Replacement="/",
                                                       LookAt=constants.xlPart,
                                                       SearchOrder=constants.xlByRows,
                                                       MatchCase=False,
                                                       SearchFormat=False,
                                                       ReplaceFormat=False
                                                       )

    Refresh_Status_label("设置<空间管理表>单元格格式......", 55.0)

    # 判断Sheet数量，计算楼层表格是否够用，并且命名Sheet
    Prepare_Floor_Sheets(Excel_Sheets, Get_Max_Row_Num(Excel_Sheets(1), 0))

    # 先在合计表中按照类别排序 或者 按照自定义顺序排序

    if sort_option == 0:
        Excel_Sheets(1).Range(All_Cells_Range).Sort(Key1=Excel_Sheets(1).Range("%s2:%s%s" % (Column_Name[Final_Table_Title.index("楼层")],
                                                                                             Column_Name[
            Final_Table_Title.index("楼层")],
            str(Excel_Sheets(
                1).UsedRange.Rows.Count)
        )),
            Key2=Excel_Sheets(1).Range("%s2:%s%s" % (Column_Name[Final_Table_Title.index("类别")],
                                                     Column_Name[
                Final_Table_Title.index("类别")],
                str(Excel_Sheets(
                    1).UsedRange.Rows.Count)
            )),
            Header=1,
            MatchCase=False,
            Orientation=constants.xlTopToBottom,
            SortMethod=1
        )
    else:
        Excel_Sheets(1).Range(All_Cells_Range).Sort(Key1=Excel_Sheets(1).Range("%s1" % Column_Name[Final_Table_Title.index("专柜")]),
                                                    Header=1,
                                                    OrderCustom=Excel_App.CustomListCount + 1,
                                                    MatchCase=False,
                                                    Orientation=constants.xlTopToBottom,
                                                    SortMethod=1
                                                    )

    Refresh_Status_label("创建各楼层工作表......", 65.0)
    # 按照要求设置每张楼层表格中的数据
    for Sheet_Num in range(1, Excel_Sheets.Count + 1):
        if "Sheet" in Excel_Sheets(Sheet_Num).Name:
            Excel_Sheets(Sheet_Num).Delete()
        else:
            Refresh_Status_label("创建<%s>层工作表......" % Excel_Sheets(
                Sheet_Num).Name, 65.0 + Sheet_Num * 5)
            Set_Range_Layout(obj=Excel_Sheets(Sheet_Num).Range(All_Cells_Range),
                             type="Range")

            Set_Range_Layout(obj=Excel_Sheets(Sheet_Num).Rows("1:1"),
                             type="Row",
                             RowHeight=39)

            if Sheet_Num != 1:

                Excel_Sheets(1).Range("$%s$1:$%s$%s" % (Column_Name[Final_Table_Title.index("楼层")],
                                                        Column_Name[
                                                            Final_Table_Title.index("楼层")],
                                                        str(Excel_Sheets(
                                                            1).UsedRange.Rows.Count)
                                                        )).AutoFilter(Field=1, Criteria1=Excel_Sheets(Sheet_Num).Name)
                # 带格式复制
                Excel_Sheets(1).Range(All_Cells_Range).SpecialCells(12).Copy()

                # http://www.cnblogs.com/hdl217/p/3494790.html
                Excel_Sheets(Sheet_Num).Range("A1").PasteSpecial(Paste=-4104,
                                                                 Operation=-4142,
                                                                 SkipBlanks=False,
                                                                 Transpose=False
                                                                 )

            # 在专柜列的最后一行 添加 合计 字样
            Excel_Sheets(Sheet_Num).Cells(Get_Max_Row_Num(Excel_Sheets(Sheet_Num), 1),
                                          Final_Table_Title.index("专柜")).Value = "合计"

            # 在每张表中添加合计行
            for Sum_column in Sum_Formula_info:
                Row_Count = Get_Max_Row_Num(Excel_Sheets(Sheet_Num), 0)
                Column_Num = Final_Table_Title.index(Sum_column)

                if Sum_Formula_info[Sum_column]["Method"] == "合计":
                    Column_1st_Char = Column_Name[Final_Table_Title.index(
                        Sum_Formula_info[Sum_column]["字段"][0])]
                    Excel_Sheets(Sheet_Num).Cells(Row_Count + 1,
                                                  Column_Num).Value = "=SUM(%s2:%s%s)" % (Column_1st_Char,
                                                                                          Column_1st_Char,
                                                                                          Row_Count)
                elif Sum_Formula_info[Sum_column]["Method"] == "比值":
                    if "元" in Sum_column:
                        Column_1st_Char = Column_Name[Final_Table_Title.index(
                            Sum_Formula_info[Sum_column]["字段"][0])]
                        Column_2rd_Char = Column_Name[Final_Table_Title.index(
                            Sum_Formula_info[Sum_column]["字段"][1])]
                        Excel_Sheets(Sheet_Num).Cells(Row_Count + 1,
                                                      Column_Num).Value = "=%s%s/%s%s*10000" % (Column_1st_Char,
                                                                                                Row_Count + 1,
                                                                                                Column_2rd_Char,
                                                                                                Row_Count + 1
                                                                                                )
                    else:
                        Column_1st_Char = Column_Name[Final_Table_Title.index(
                            Sum_Formula_info[Sum_column]["字段"][0])]
                        Column_2rd_Char = Column_Name[Final_Table_Title.index(
                            Sum_Formula_info[Sum_column]["字段"][1])]
                        Excel_Sheets(Sheet_Num).Cells(Row_Count + 1,
                                                      Column_Num).Value = "=%s%s/%s%s-1" % (Column_1st_Char,
                                                                                          Row_Count + 1,
                                                                                          Column_2rd_Char,
                                                                                          Row_Count + 1
                                                                                          )

                else:
                    Column_1st_Char = Column_Name[Final_Table_Title.index(
                        Sum_Formula_info[Sum_column]["字段"][0])]
                    Column_2rd_Char = Column_Name[Final_Table_Title.index(
                        Sum_Formula_info[Sum_column]["字段"][1])]
                    Excel_Sheets(Sheet_Num).Cells(Row_Count + 1,
                                                  Column_Num).Value = "=%s%s/%s%s-1" % (Column_1st_Char,
                                                                                        Row_Count + 1,
                                                                                        Column_2rd_Char,
                                                                                        Row_Count + 1
                                                                                        )
            # 使用复制粘贴值，去除合计行中的公式引用
            Excel_Sheets(Sheet_Num).Range(All_Cells_Range).Copy()
            Excel_Sheets(Sheet_Num).Range(All_Cells_Range).PasteSpecial(Paste=-4163,
                                                                        Operation=-4142,
                                                                        SkipBlanks=False,
                                                                        Transpose=False
                                                                        )
    Refresh_Status_label("核对表格数据......", 93.0)
    # 取消 Sheet1 合计表的 筛选状态
    Excel_Sheets(1).Range("$%s$1:$%s$%s" % (Column_Name[Final_Table_Title.index("楼层")],
                                            Column_Name[
                                                Final_Table_Title.index("楼层")],
                                            str(Excel_Sheets(
                                                1).UsedRange.Rows.Count)
                                            )).AutoFilter(Field=1)

    for Sheet_Num in range(1, Excel_Sheets.Count + 1):
        Set_Range_Layout(obj=Excel_Sheets(Sheet_Num).Range(
            "J:J,N:N,Q:Q,V:V,W:W,X:X,AA:AA,AB:AB,AC:AC,AD:AD,AG:AG,AH:AH,AI:AI,AJ:AJ,AK:AK,AL:AL,AM:AM"), type="Column", NumberFormatLocal="0.00%")
        if del_option != 0:
            Refresh_Status_label("删除同期相关数据......", 95.0)
            Excel_Sheets(Sheet_Num).Range(
                "A:A,B:B,C:C,P:P,U:U,Z:Z,AC:AC,AD:AD,AF:AF,AG:AG,AI:AI,AJ:AJ,AN:AN,AO:AO").Delete(Shift=constants.xlToLeft)

    Refresh_Status_label("正在保存......", 97.0)
    Excel_Workbook.SaveAs(FilePath, FileFormat=51)
    Refresh_Status_label("已经完成......", 100.0)
    # 打开空间管理表
    if sort_option == 0:
        if tkinter.messagebox.askyesno("提示！", "空间管理表已经完成！是否立即打开？"):
            Excel_App = Open_Excel_By_Win32com(True)
            Excel_Workbook = Open_Workbook_By_Win32com(Excel_App, FilePath)
        else:
            Excel_Workbook.Close()
    else:
        tkinter.messagebox.showinfo("提示！", "勾选'自定义排序'后请手动打开文件，点击‘是’修复即可！")
        Excel_Workbook.Close()


def Check_registration_Status_label(ip, filename, keyvalue):
    global isRegistered, UserName, Company, Department
    Registration = ckr.registration_check(ip, filename, keyvalue)
    if Registration[0]:
        isRegistered = True
        UserName = Registration[1]["UserName"]
        Company = Registration[1]["Company"]
        Department = Registration[1]["Department"]

        Refresh_Status_label(("...已激活..."), 0)
    else:
        isRegistered = False
        Refresh_Status_label(("...未激活..."), 0)


if __name__ == '__main__':

    loadview()
